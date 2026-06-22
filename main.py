from playwright.sync_api import sync_playwright
from login_siafe import fazer_login
from extrair_relatorio import baixar_relatorio
import os
from dotenv import load_dotenv
import pandas as pd
from datetime import datetime
import pandas as pd
import cruzamento
import extrair_dados
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(
    level=logging.INFO,
    filename="app.log",
    filemode="a",
    format="%(asctime)s - %(levelname)s - %(message)s"
)

logging.info("Aplicação iniciou")

# ==================================

# DEFINE NOME DO ARQUIVO E PASTA QUE SERÁ SALVO

# ==================================

date_now = datetime.now().strftime('%d.%m.%Y')
ano = datetime.now().strftime('%Y')
arquivo = f'ARQUIVOS_RETORNO/{date_now}_PROCESSOS_EMPENHADOS_LIQUIDADOS_PAGOS {ano}.xlsx'

# ==================================

# LOGANDO NO .ENV

# ==================================

load_dotenv()

CPF = os.getenv("USUARIO")
SENHA = os.getenv("SENHA")
URL = os.getenv("LINK")

# ==================================

# FUNÇÃO PARA EXTRAIR OS RELATÓRIOS NECESSÁRIOS DO FLEXVISION

# ==================================

def extraindo_relatorios_flexvision(
    usuario="Cainã",
    codigo="082831",
    arquivo="relatorio.xlsx"
    ):
    
    with sync_playwright() as p:
        browser, context, page = fazer_login(
            p,
            CPF,
            SENHA
        )

        baixar_relatorio(
            page,
            usuario=usuario,
            codigo=codigo,
            arquivo=arquivo
        )
        browser.close()

# ==================================

# FAZENDO RELATÓRIO

# ==================================

relatorios_flexvision = {'082831':'data/Relatório_PDs.xls',
                         '083008':'data/Relatório_OB.xls',
                         '083079':'data/Relatório_Lançamentos Contábeis.xls',
                         '083098':'data/Relatório NE.xls',
                         '083100':'data/Relatório - NL.xls',
                         '082836':'data/OB _ Apenas data.xls'}
for cod in relatorios_flexvision:
    extraindo_relatorios_flexvision(    
        codigo=cod,
        arquivo=relatorios_flexvision[cod]
        )

# Abre arquivos de base
de_para = pd.read_excel('config/De_para.xlsx')
df_layout = pd.read_excel('config/Layout.xlsx')

# Define variáveis DF
df_pd, df_ne, df_nl, df_lancamentos = cruzamento.dados('data')
df_geral, df_filtrado = extrair_dados.relatorio(df_pd)
df_ne_nl = extrair_dados.ne_nl(df_ne, df_nl, df_lancamentos)

# Cruza dados
cruzamento.cruzar_dados(de_para,'Para','Contrato',df_ne_nl,'Credor','Contrato')
cruzamento.cruzar_dados(de_para,'Para','Contrato',df_pd,'Credor','Contrato')
cruzamento.cruzar_dados(de_para,'Para','Contrato',df_geral,'CREDOR','Contrato')

cruzamento.cruzar_dados(df_ne_nl,'Contrato','NE',df_layout,'CONTRATO','EMPENHOS (NE)')
cruzamento.cruzar_dados(df_ne_nl,'Contrato','NL',df_layout,'CONTRATO','LIQUIDAÇÕES (NL)')

# Lista dos credores - Empresas
credores_list = ['GRUPO SBV', 'JRV EMPREENDIMENTOS', 'BELLAMAR', 'G. S. GOUVEA',
    'CONPLAN', 'INVICTA COMERCIAL', 'F P VIEIRA ENGENHARIA', 'CONSTRUMAX', 
    '4X4 TERRAPLANAGEM', 'FARAÓ', 'GUIMAVE', 'CONSTRUSAN', 'FERDAN']

# Filtra apenas o que não houve retorno do banco
list_enviado = [
    'Aguardando Envio',
    'Enviado Manualmente',
    'Enviado ao Banco'
]


# ====================================================
# VERIFICA DATAS QUE AINDA NÃO HOUVE RETORNO DO BANCO
# ====================================================

# filtra
df_filtrado_enviado = df_pd[df_pd['Situação envio'].isin(list_enviado)]

# filtra apenas os credores desejados
df_filtrado = df_filtrado_enviado[
    df_filtrado_enviado['Credor'].isin(credores_list)
]

# cria pivot
df_credor_datas = (
    df_filtrado
    .pivot_table(
        index='Credor',
        columns='Data Emissão',
        values='2026',
        aggfunc='sum',
        fill_value=0
    )
)

# remove colunas sem valores
df_credor_datas = df_credor_datas.loc[
    :,
    (df_credor_datas != 0).any(axis=0)
]

# volta índice para coluna
df_credor_datas = df_credor_datas.reset_index()

# Pega apenas as colunas que serão adicionadas
novas_colunas = df_credor_datas.columns[1:]

# Faz o merge alinhando Credor -> CREDOR
df_geral = df_geral.merge(
    df_credor_datas[['Credor', *novas_colunas]],
    left_on='CREDOR',
    right_on='Credor',
    how='left'
)

# Remove coluna duplicada criada pelo merge
df_geral.drop(columns='Credor', inplace=True)

# Lista as colunas que serão anexadas do df_geral
colunas_datas = list(df_credor_datas.columns[1:])

# Cruza dados para adicionar a nova coluna
for coluna in colunas_datas:
    coluna_referencia = 'OBs PROCESSADAS E PAGAS'

    cols = df_geral.columns.tolist()

    # remove a nova coluna da posição atual
    cols.remove(coluna)

    # encontra posição da referência
    idx = cols.index(coluna_referencia)

    # insere na posição desejada
    cols.insert(idx, coluna)

    # reordena dataframe
    df_geral = df_geral[cols]

# ====================================================
# RENOMEIA E REPOSICIONA AS COLUNAS QUE FORAM ADICIONADAS
# ====================================================

col_inicio = 'OBs ENVIADAS AO BANCO A SEREM PROCESSADAS E PAGAS'
col_fim = 'OBs PROCESSADAS E PAGAS'

# lista de colunas
cols = df_geral.columns.tolist()

# posições
idx_inicio = cols.index(col_inicio)
idx_fim = cols.index(col_fim)

# colunas entre elas
colunas_renomear = cols[idx_inicio + 1:idx_fim]

# renomeia
df_todas_datas = df_geral.rename(columns={
    col: f'{col.strftime('%d-%m-%Y')} - OBs ENVIADAS AO BANCO'
    for col in colunas_renomear
})

# lista de colunas
cols = df_todas_datas.columns.tolist()

# posições
idx_inicio = cols.index(col_inicio)
idx_fim = cols.index(col_fim)

# colunas entre elas
colunas_renomeadas = cols[idx_inicio + 1:idx_fim]

# ====================================================
# PADRONIZA OS DFS COM O LAYOUT PARA EXPORTAÇÃO
# ====================================================

# Cruzar dados para descobrir Situação envio
mapa_situacao = (
    df_pd
    .drop_duplicates('Credor')
    .set_index('Credor')['Contrato']
)

df_geral['Contrato'] = df_geral['CREDOR'].map(mapa_situacao)

mapa_situacao = (
    df_geral
    .drop_duplicates('Contrato')
    .set_index('Contrato')[
        [
            'PRESTADORES DE SERVIÇOS QUE FALTAM EMITIR PD',
            'RETENÇÕES QUE FALTAM EMITIR PD',
            'PROGRAMAÇÃO DE DESEMBOLSO',
            'PDs QUE FALTAM SER EXECUTADAS',
            'REGISTRO DE ENVIO (RE) DAS ORDENS BANCÁRIAS (OB)',
        ]
    ]
    .rename(columns={
        'PROGRAMAÇÃO DE DESEMBOLSO': 'PROGRAMAÇÃO DE DESEMBOLSO (PD)'
    })
)

df_layout = df_layout.join(
    mapa_situacao,
    on='CONTRATO'
)

# Criar numeração começando em 1
df_layout.insert(0, "N°", range(1, len(df_layout) + 1))

# Cria coluna 'SALDO DOS EMPENHOS POR CONTRATO'
df_layout['SALDO DOS EMPENHOS POR CONTRATO'] = None

# Nomeia as abas em um df
df_aba_1 = df_layout.copy()
df_aba_2 = df_layout[df_layout['CREDOR'].isin(credores_list)]

df_aba_3 = df_todas_datas[df_todas_datas['CREDOR'].isin(credores_list)]
df_aba_3 = df_aba_3.drop(columns=[
    'PRESTADORES DE SERVIÇOS QUE FALTAM EMITIR PD',
    'RETENÇÕES QUE FALTAM EMITIR PD',
    'PROGRAMAÇÃO DE DESEMBOLSO',
    'PDs QUE FALTAM SER EXECUTADAS',
    'Contrato'
])

DF_BASE_PD = df_pd.drop(columns=['Documento Alterado NE','Status execução.1','2026'])

# ====================================================
# EXPORTA OS DATA FRAMES COM O LAYOUT PADRÃO    
# ====================================================

with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:

    # ==========================================
    # EXPORTA DF
    # ==========================================

    df_aba_1.to_excel(
        writer,
        sheet_name='Execução_NE_NL_PD_OB',
        index=False,
        startrow=2
    )

    # ==========================================    
    # ACESSA PLANILHA
    # ==========================================

    wb = writer.book
    ws = writer.sheets['Execução_NE_NL_PD_OB']

    # ==========================================
    # TÍTULO
    # ==========================================

    ws.merge_cells('A1:M1')

    ws['A1'] = (
        'EXECUÇÃO ORÇAMENTÁRIA / EMPENHOS - '
        'LIQUIDAÇÕES - PAGAMENTOS / SEDIPAF 2026'
    )

    ws['A1'].font = Font(
        bold=True,
        size=16
    )

    ws['A1'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # ==========================================
    # ESTILOS
    # ==========================================

    fill_cabecalho = PatternFill(
        start_color='D9D9D9',
        end_color='D9D9D9',
        fill_type='solid'
    )

    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==========================================
    # CABEÇALHO
    # ==========================================

    for cell in ws[3]:

        cell.font = Font(bold=True)

        cell.fill = fill_cabecalho

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        cell.border = borda

    # ==========================================
    # FORMATO MOEDA
    # ==========================================

    colunas_moeda = ['F','G','H','I','J','K','L']

    for col in colunas_moeda:

        for cell in ws[col]:

            if cell.row > 3:
                cell.number_format = 'R$ #,##0.00'

    # ==========================================
    # LARGURA COLUNAS
    # ==========================================

    larguras = {
        'A': 8,
        'B': 30,
        'C': 14,
        'D': 35,
        'E': 40,
        'F': 18,
        'G': 18,
        'H': 18,
        'I': 18,
        'J': 18,
        'K': 18,
        'L': 18,
        'M': 18
    }

    for col, largura in larguras.items():

        ws.column_dimensions[col].width = largura

    # ==========================================
    # LINHAS AMARELAS
    # ==========================================

    amarelo = PatternFill(
        start_color='FFFF00',
        end_color='FFFF00',
        fill_type='solid'
    )

    for row in ws.iter_rows(min_row=4):

        credor = row[3].value
        
        if str(credor).upper() in credores_list:
            for cell in row:
                cell.fill = amarelo
                cell.border = borda

    # ==========================================
    # BORDAS GERAIS
    # ==========================================

    for row in ws.iter_rows():

        for cell in row:
            cell.border = borda

    # ==========================================
    # ADICIONAR FORMULA
    # ==========================================

    for row in range(4, ws.max_row + 1):

        cell = ws[f'M{row}']

        cell.value = f'=F{row}-G{row}'
        cell.number_format = 'R$ #,##0.00'

    for row in range(4, ws.max_row + 1):

        cell = ws[f'M{row}']

        cell.value = f'=F{row}-G{row}'
        cell.number_format = 'R$ #,##0.00'


    linha_total = ws.max_row + 1

    # Mesclar A:E
    ws.merge_cells(f'A{linha_total}:E{linha_total}')

    # Escrever TOTAL
    ws[f'A{linha_total}'] = 'TOTAL'

    # Estilo
    ws[f'A{linha_total}'].font = Font(bold=True)

    ws[f'A{linha_total}'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # ================
    # TOTAL
    # ================

    fill_header = PatternFill(
        start_color='D9D9D9',
        end_color='D9D9D9',
        fill_type='solid'
    )

    thin = Side(style='thin')

    ws[f'A{linha_total}'].fill = fill_header

    ws[f'A{linha_total}'].border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )
    
    for col in range(6, 14):

        letra = get_column_letter(col)

        ws[f'{letra}{linha_total}'] = (
            f'=SUM({letra}4:{letra}{linha_total-1})'
        )

        ws[f'{letra}{linha_total}'].font = Font(bold=True)

        ws[f'{letra}{linha_total}'].number_format = 'R$ #,##0.00'

    # =========================================================

    # FORMATANDO ABA 2

    # =========================================================

    # começa na linha 2
    df_aba_2.to_excel(writer,
                    sheet_name='NE_NL_PD_OB_13 Empresa',
                    startrow=2,
                    index=False)

    wb = writer.book
    ws = writer.sheets['NE_NL_PD_OB_13 Empresa']

    # =========================
    # TÍTULO
    # =========================

    ws.merge_cells('A1:M1')

    ws['A1'] = (
        'EXECUÇÃO ORÇAMENTÁRIA / EMPENHOS - '
        'LIQUIDAÇÕES - PAGAMENTOS / SEDIPAF 2026 (26/05/26)'
    )

    ws['A1'].font = Font(
        bold=True,
        size=14
    )

    ws['A1'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # =========================
    # ESTILO CABEÇALHO
    # =========================

    fill_header = PatternFill(
        start_color='D9D9D9',
        end_color='D9D9D9',
        fill_type='solid'
    )

    thin = Side(style='thin')

    for cell in ws[3]:

        cell.font = Font(bold=True)

        cell.fill = fill_header

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

    # =========================
    # FORMATAÇÃO DAS CÉLULAS
    # =========================

    colunas_monetarias = [
        'F', 'G', 'H', 'I',
        'J', 'K', 'L', 'M'
    ]

    for col in colunas_monetarias:

        for cell in ws[col][3:]:

            cell.number_format = 'R$ #,##0.00'

    # =========================
    # BORDAS E ALINHAMENTO
    # =========================

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row,
        min_col=1,
        max_col=13
    ):

        for cell in row:

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            cell.alignment = Alignment(
                vertical='center'
            )

    # =========================
    # LARGURA DAS COLUNAS
    # =========================

    larguras = {
        'A': 8,
        'B': 30,
        'C': 12,
        'D': 28,
        'E': 35,
        'F': 18,
        'G': 18,
        'H': 18,
        'I': 18,
        'J': 20,
        'K': 18,
        'L': 18,
        'M': 20,
    }

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    # =========================
    # ALTURA DO CABEÇALHO
    # =========================

    ws.row_dimensions[3].height = 40

    # =========================
    # SALDO DOS CONTRATOS
    # =========================
    for row in range(4, ws.max_row + 1):

        cell = ws[f'M{row}']

        cell.value = f'=F{row}-G{row}'
        cell.number_format = 'R$ #,##0.00'

    # =========================
    # LINHA TOTAL
    # =========================
    
    linha_total = ws.max_row + 1

    # Mesclar A:E
    ws.merge_cells(f'A{linha_total}:E{linha_total}')

    # Escrever TOTAL
    ws[f'A{linha_total}'] = 'TOTAL'

    # Estilo
    ws[f'A{linha_total}'].font = Font(bold=True)

    ws[f'A{linha_total}'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    ws[f'A{linha_total}'].fill = fill_header

    ws[f'A{linha_total}'].border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for col in range(6, 14):

        letra = get_column_letter(col)

        ws[f'{letra}{linha_total}'] = (
            f'=SUM({letra}4:{letra}{linha_total-1})'
        )

        ws[f'{letra}{linha_total}'].font = Font(bold=True)

        ws[f'{letra}{linha_total}'].number_format = 'R$ #,##0.00'

    # =========================
    # CONGELAR PAINEL
    # =========================

    ws.freeze_panes = 'A4'

    # =========================================================

    # FORMATANDO ABA 3

    # =========================================================
    # =========================================================
    # ABA 3 - RELAÇÃO DE ENVIO
    # =========================================================

    df_aba_3.to_excel(
        writer,
        sheet_name='NE_NL_PD_OB_Deodônio',
        index=False,
        startrow=2,
        header=False
    )

    # =========================================
    # ACESSA ABA
    # =========================================

    ws = writer.sheets['NE_NL_PD_OB_Deodônio']

    # =========================================
    # COLUNAS
    # =========================================

    colunas_fixas_1 = [
        'CREDOR',
        'REGISTRO DE ENVIO (RE) DAS ORDENS BANCÁRIAS (OB)',
        'OBs ENVIADAS AO BANCO A SEREM PROCESSADAS E PAGAS',
    ]

    colunas_fixas_2 = [
        'OBs PROCESSADAS E PAGAS',
        '>MES ATUAL',
        '>7 DIAS'
    ]

    colunas_datas = [
        c for c in df_aba_3.columns
        if c not in colunas_fixas_1 + colunas_renomeadas + colunas_fixas_2
    ]

    colunas_finais = (
        colunas_fixas_1 +
        colunas_renomeadas +
        colunas_fixas_2+
        colunas_datas
    )

    ultima_coluna = len(colunas_finais)

    # =========================================
    # ESTILOS
    # =========================================

    fill_cinza = PatternFill(
        start_color='D9D9D9',
        end_color='D9D9D9',
        fill_type='solid'
    )

    fill_verde = PatternFill(
        start_color='E2EFDA',
        end_color='E2EFDA',
        fill_type='solid'
    )

    fill_azul = PatternFill(
        start_color='DDEBF7',
        end_color='DDEBF7',
        fill_type='solid'
    )

    fill_vermelho = PatternFill(
        start_color='EAD1D1',
        end_color='EAD1D1',
        fill_type='solid'
    )

    thin = Side(style='thin')

    borda = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=ultima_coluna
    )

    titulo = ws.cell(row=1, column=1)

    titulo.value = (
        'RELAÇÃO DE ENVIO (RE) / ORDEM BANCÁRIA (OB) '
        '/ SEDIPAF 2026 (26/05/26)'
    )

    titulo.font = Font(
        bold=True,
        size=13
    )

    titulo.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # =========================================
    # CABEÇALHO
    # =========================================

    for idx, coluna in enumerate(colunas_finais, start=1):

        cell = ws.cell(
            row=3,
            column=idx
        )

        cell.value = coluna

        cell.font = Font(
            bold=True,
            size=9
        )

        cell.border = borda

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        # cores
        if coluna in colunas_fixas_1 or colunas_fixas_2[0]:

            cell.fill = fill_cinza

        elif coluna in colunas_renomeadas:

            cell.fill = fill_verde

        elif coluna in colunas_datas or colunas_fixas_2[1:]:

            cell.fill = fill_azul

        else:

            cell.fill = fill_cinza

    # =========================================
    # FORMATA DADOS
    # =========================================

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row,
        min_col=1,
        max_col=ultima_coluna
    ):

        for cell in row:

            cell.border = borda

            cell.alignment = Alignment(
                vertical='center'
            )

            # moeda
            if cell.column != 1 and cell.column != ultima_coluna+1:

                cell.number_format = 'R$ #,##0.00'

    # =========================================
    # COLORIR COLUNAS
    # =========================================

    for idx, coluna in enumerate(colunas_finais, start=1):

        if coluna in colunas_renomeadas:

            for row in range(4, ws.max_row + 1):

                ws.cell(row=row, column=idx).fill = fill_verde

        elif coluna in colunas_datas or coluna in ['>MES ATUAL', '>7 DIAS']:

            for row in range(4, ws.max_row + 1):

                ws.cell(row=row, column=idx).fill = fill_azul

        elif coluna in ['OBs ENVIADAS AO BANCO A SEREM PROCESSADAS E PAGAS', 'OBs PROCESSADAS E PAGAS']:

            for row in range(4, ws.max_row + 1):

                ws.cell(row=row, column=idx).fill = fill_vermelho

    # =========================================
    # TOTAL
    # =========================================

    linha_total = ws.max_row + 1

    ws.merge_cells(
        start_row=linha_total,
        start_column=1,
        end_row=linha_total,
        end_column=1
    )

    cell_total = ws.cell(
        row=linha_total,
        column=1
    )

    cell_total.value = 'TOTAL'

    cell_total.font = Font(bold=True)

    cell_total.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    cell_total.fill = fill_cinza

    cell_total.border = borda

    # SOMATÓRIOS
    for col in range(2, ultima_coluna+1):

        letra = get_column_letter(col)

        cell = ws.cell(
            row=linha_total,
            column=col
        )

        cell.value = (
            f'=SUM({letra}4:{letra}{linha_total-1})'
        )

        cell.font = Font(bold=True)

        cell.number_format = 'R$ #,##0.00'

        cell.border = borda

    # =========================================
    # LARGURAS
    # =========================================

    larguras = {
        'A': 30,
        'B': 22,
        'C': 20,
        'D': 20,
        'E': 18,
        'F': 16,
        'G': 16,
    }

    for col, largura in larguras.items():

        ws.column_dimensions[col].width = largura

    # datas
    for col in range(8, ultima_coluna + 1):

        letra = get_column_letter(col)

        ws.column_dimensions[letra].width = 14

    # =========================================
    # ALTURAS
    # =========================================

    ws.row_dimensions[3].height = 55

    # =========================================
    # CONGELAR
    # =========================================

    ws.freeze_panes = 'A4'

    DF_BASE_PD.to_excel(writer,
                   sheet_name='analítico_pd',
                   index=False)
    
    df_ne.to_excel(writer,
                   sheet_name='analítico_ne',
                   index=False)
    
    df_nl.to_excel(writer,
                   sheet_name='analítico_ne',
                   index=False)
    
logging.info(f'Arquivo salvo: {arquivo}')